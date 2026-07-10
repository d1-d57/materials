# =====================================================================
# Модель ценообразования каршеринга (Москва). Генерирует carsharing_pricing.xlsx.
#
# ИСТОЧНИК ПРАВДЫ ПО ВХОДАМ: carsharing_data_sources.md (реестр + changelog).
# Приватных данных оператора нет — всё из задачи (📋) и открытых данных Москвы (📊).
#
# ФАЗА 1 (заземление на открытые данные Москвы), внесено в этот файл:
#   • B6  утилизация минутной базы 0.30→0.18  — 📊 эконом ~94-110 км/сут (Дептранс/Трушеринг)
#                                                 ÷ эфф.скорость ÷ 24ч → ~16-21%; 30% это среднее
#                                                 по парку, завышено пакетами.
#   • B8  скорость 25→24, переописана как ЭФФЕКТИВНАЯ (дистанция/длительность) — 📊 ср.поездка
#                                                 Москва 23км/63мин≈22; Делимобиль 24.4-26.6 (B1).
#   • u25 предельный пол: убран ×ходовая доля (B11) — был ДВОЙНОЙ УЧЁТ: эффективная скорость
#                                                 простой внутри поездки уже включает. Пол=15×24/60=6.0.
#   • B27 рыночный ориентир 7.5→8.5 — 📊 Москва эконом «от 8.5 ₽/мин» (Т—Ж 2024).
#   Результат: пробег 180→104 км (совпал с наблюдаемым ✓), средний таргет 15.5→~19.6 (ВЫШЕ),
#   три уровня: пол 6 < рынок 8.5 ≪ таргет 20. Подробности и что это перевернуло — в changelog реестра.
# =====================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

BLUE = Font(name='Arial', color='0000FF')
BLACK = Font(name='Arial', color='000000')
GREEN = Font(name='Arial', color='008000')
BOLD = Font(name='Arial', bold=True)
BOLDW = Font(name='Arial', bold=True, color='FFFFFF')
TITLE = Font(name='Arial', bold=True, size=13)
YELLOW = PatternFill('solid', start_color='FFFF00')
HEADER = PatternFill('solid', start_color='305496')
GREY = PatternFill('solid', start_color='D9D9D9')
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
RUB = '#,##0'
RUB1 = '#,##0.0'
PCT = '0%'
PCT1 = '0.0%'

wb = Workbook()

# ===================== ПАРАМЕТРЫ =====================
p = wb.active
p.title = 'Параметры'
p['A1'] = 'Параметры модели — меняем синие ячейки на жёлтом'
p['A1'].font = TITLE
for c, t in [('A3','Параметр'),('B3','Значение'),('C3','Ед.'),('D3','Комментарий')]:
    p[c] = t; p[c].font = BOLDW; p[c].fill = HEADER; p[c].border = BORDER

def inp(cell, val, unit, comment, fmt=RUB):
    r = cell[1:]
    p[f'A{r}'] = comment_label[cell]; p[f'A{r}'].font = BLACK
    p[cell] = val; p[cell].font = BLUE; p[cell].fill = YELLOW; p[cell].number_format = fmt; p[cell].border = BORDER
    p[f'C{r}'] = unit; p[f'C{r}'].font = BLACK
    p[f'D{r}'] = comment; p[f'D{r}'].font = Font(name='Arial', italic=True, size=9, color='808080')

comment_label = {
 'B4':'Фиксированные затраты (лизинг + страховка)',
 'B5':'Переменные затраты на пробег',
 'B6':'Утилизация минутной базы (эконом, доля времени под заказом)',
 'B7':'Целевая GP-маржа',
 'B8':'Эффективная скорость (дистанция/длительность поездки)',
 'B9':'Плата за км в пакетах',
 'B10':'Поездок на машину в сутки',
 'B11':'Ходовая доля (R4): доля времени аренды в движении',
 'B12':'Теневая цена ёмкости в пик (R5)',
}
inp('B4', 2000, 'руб/сут', 'дано')
inp('B5', 15, 'руб/км', 'дано')
inp('B6', 0.18, 'доля', '\U0001F4CA Москва: эконом ~94-110 км/сут / эфф.скорость -> util ~16-21%; задача даёт 30% по парку, но это среднее, завышенное пакетами', PCT)
inp('B7', 0.30, 'доля', 'цель по gross profit', PCT)
inp('B8', 24, 'км/ч', '\U0001F4CA Москва: ср.поездка 23 км/63 мин = 22; Делимобиль 24-27. Эффективная (со стоянками внутри) — НЕ скорость в движении', RUB)
inp('B9', 18, 'руб/км', 'выше предельных издержек по условию Шмалензее (R6: предельный участник легче среднего); коридор [15;21]', RUB)
inp('B10', 10, 'шт/сут', '≈ 432 мин под заказом / ~40 мин на поездку', RUB)
inp('B11', 0.55, 'доля', 'справочно (Sprei): в эффективную скорость B8 уже заложена; для предельного пола НЕ используется — иначе двойной учёт', PCT)
inp('B12', 0, 'руб/мин', 'R5: off-peak ≈ 0; в пик растёт и поднимает предельный пол (peak-load)', RUB1)

p['A13'] = 'Утилизация по длительности аренды (доля времени под заказом)'
p['A13'].font = BOLD
p['A14'] = 'Тариф'; p['B14'] = 'Утилизация'
for c in ('A14','B14'):
    p[c].font = BOLDW; p[c].fill = HEADER; p[c].border = BORDER
p['D14'] = 'ВАЖНО: замените на реальную утилизацию по длительностям из вашей статистики'
p['D14'].font = Font(name='Arial', italic=True, size=9, color='C00000')
util = [('1 час',0.40),('2 часа',0.45),('3 часа',0.50),('6 часов',0.55),
        ('12 часов',0.70),('1 сутки',0.85),('7 суток',0.95)]
r = 15
for name, u in util:
    p[f'A{r}'] = name; p[f'A{r}'].font = BLACK; p[f'A{r}'].border = BORDER
    p[f'B{r}'] = u; p[f'B{r}'].font = BLUE; p[f'B{r}'].fill = YELLOW; p[f'B{r}'].number_format = PCT; p[f'B{r}'].border = BORDER
    r += 1

p.column_dimensions['A'].width = 48
p.column_dimensions['B'].width = 12
p.column_dimensions['C'].width = 9
p.column_dimensions['D'].width = 60

# ===================== ЮНИТ-ЭКОНОМИКА =====================
u = wb.create_sheet('Юнит-экономика')
u['A1'] = 'Юнит-экономика средней машины (база = минутный тариф)'
u['A1'].font = TITLE

def row(sh, r, label, formula, unit, fmt=RUB, font=BLACK, bold=False):
    sh[f'A{r}'] = label; sh[f'A{r}'].font = BOLD if bold else BLACK
    sh[f'B{r}'] = formula; sh[f'B{r}'].font = font; sh[f'B{r}'].number_format = fmt; sh[f'B{r}'].border = BORDER
    sh[f'C{r}'] = unit; sh[f'C{r}'].font = Font(name='Arial', size=9, color='808080')

row(u,3,'Часов под заказом в сутки', "=24*Параметры!B6", 'ч', RUB1, GREEN)
row(u,4,'Минут под заказом в сутки', "=B3*60", 'мин', RUB)
row(u,5,'Пробег в сутки', "=B3*Параметры!B8", 'км', RUB, GREEN)
row(u,7,'Фиксированные затраты / сут', "=Параметры!B4", 'руб', RUB, GREEN)
row(u,8,'Переменные затраты / сут', "=B5*Параметры!B5", 'руб', RUB, GREEN)
row(u,9,'Итого затраты / сут', "=B7+B8", 'руб', RUB, BLACK, True)
row(u,11,'Выручка для целевой маржи', "=B9/(1-Параметры!B7)", 'руб', RUB, GREEN, True)
row(u,12,'Gross Profit / сут', "=B11-B9", 'руб', RUB)
row(u,13,'Факт. GP-маржа (проверка)', "=B12/B11", '%', PCT1)
u['A15'] = 'Минутный тариф'; u['A15'].font = BOLD
row(u,16,'Ставка, руб/мин', "=B11/B4", 'руб/мин', RUB1, BLACK, True)
row(u,17,'— фикс. компонента (покрывает простой)', "=B7/B4", 'руб/мин', RUB1)
row(u,18,'— перем. компонента (себест. за минуты движения)', "=(Параметры!B8/60)*Параметры!B5", 'руб/мин', RUB1, GREEN)
row(u,19,'— себестоимость, руб/мин', "=B17+B18", 'руб/мин', RUB1)
row(u,20,'— с маржой /(1-маржа), сверка с B16', "=B19/(1-Параметры!B7)", 'руб/мин', RUB1, GREEN)
u['B16'].fill = PatternFill('solid', start_color='FFF2CC')

# ---- R5: предельный пол vs средний таргет ----
u['A22'] = 'Предельный пол vs средний таргет (R5)'; u['A22'].font = BOLD
row(u,23,'Средний таргет безубыточности (FDC + маржа) = ставка B16', "=B16", 'руб/мин', RUB1, BLACK)
row(u,24,'Предельный пол — вариативная компонента, ₽ за минуту аренды', "=B18", 'руб/мин', RUB1, GREEN)
row(u,25,'Предельный пол, off-peak — на забронированную минуту (B8 эффективная -> без ×ходовая доля)', "=B18", 'руб/мин', RUB1, GREEN)
row(u,26,'Предельный пол, peak (+ теневая цена ёмкости)', "=B25+Параметры!B12", 'руб/мин', RUB1, BLACK)
row(u,27,'Рыночный ориентир эконом-минуты (\U0001F4CA Москва ~8.5)', 8.5, 'руб/мин', RUB1, BLUE)
u['B27'].fill = YELLOW
u['B24'].fill = PatternFill('solid', start_color='E2EFDA')
u['B25'].fill = PatternFill('solid', start_color='E2EFDA')
u['A28'] = 'Данные Москвы: пол ~6 < рынок ~8.5 ≪ таргет ~20. Рынок впритык над полом, далеко под таргетом — фикс отбивается не минутной ставкой'
u['A28'].font = Font(name='Arial', italic=True, size=9, color='C00000')
u['A29'] = 'Пол = операционные предельные + теневая цена ёмкости (статически peak-load, динамически Беллман)'
u['A29'].font = Font(name='Arial', italic=True, size=9, color='808080')

u.column_dimensions['A'].width = 48
u.column_dimensions['B'].width = 14
u.column_dimensions['C'].width = 9

# ===================== ТАРИФЫ =====================
t = wb.create_sheet('Тарифы')
t['A1'] = 'Тарифная сетка: фикс-ставка за час падает с длительностью — это и есть «пакет дешевле»'
t['A1'].font = TITLE
heads = ['Тариф','Утилизация','Фикс-затрата\nна час под заказом, руб',
         'Фикс-плата за час\n(с маржой), руб/ч','Часов\nв аренде',
         'Фикс-плата за\nвесь пакет, руб','Плата за км,\nруб','Себест.\nза км, руб',
         'Пример: пакет\n+ 30 км, руб']
for i, h in enumerate(heads):
    c = t.cell(row=3, column=1+i, value=h)
    c.font = BOLDW; c.fill = HEADER; c.border = BORDER
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

# rows: name, util_ref, hours
rows = [
 ('Минутный (справочно)', 'Параметры!B6', None),
 ('1 час', 'Параметры!B15', 1),
 ('2 часа', 'Параметры!B16', 2),
 ('3 часа', 'Параметры!B17', 3),
 ('6 часов', 'Параметры!B18', 6),
 ('12 часов', 'Параметры!B19', 12),
 ('1 сутки', 'Параметры!B20', 24),
 ('7 суток', 'Параметры!B21', 168),
]
r = 4
for name, uref, hours in rows:
    t[f'A{r}'] = name; t[f'A{r}'].font = BLACK; t[f'A{r}'].border = BORDER
    t[f'B{r}'] = f"={uref}"; t[f'B{r}'].font = GREEN; t[f'B{r}'].number_format = PCT; t[f'B{r}'].border = BORDER
    t[f'C{r}'] = f"=Параметры!B4/(24*B{r})"; t[f'C{r}'].font = BLACK; t[f'C{r}'].number_format = RUB; t[f'C{r}'].border = BORDER
    t[f'D{r}'] = f"=C{r}/(1-Параметры!B7)"; t[f'D{r}'].font = BLACK; t[f'D{r}'].number_format = RUB; t[f'D{r}'].border = BORDER
    if hours is None:
        for col in ('E','F','I'):
            t[f'{col}{r}'] = '—'; t[f'{col}{r}'].font = BLACK; t[f'{col}{r}'].border = BORDER; t[f'{col}{r}'].alignment = Alignment(horizontal='center')
        t[f'G{r}'] = "=(Параметры!B8/60)*Параметры!B5/(1-Параметры!B7)"; t[f'G{r}'].font = BLACK; t[f'G{r}'].number_format = RUB; t[f'G{r}'].border = BORDER
    else:
        t[f'E{r}'] = hours; t[f'E{r}'].font = BLACK; t[f'E{r}'].number_format = RUB; t[f'E{r}'].border = BORDER
        t[f'F{r}'] = f"=D{r}*E{r}"; t[f'F{r}'].font = BLACK; t[f'F{r}'].number_format = RUB; t[f'F{r}'].border = BORDER
        t[f'G{r}'] = "=Параметры!B9"; t[f'G{r}'].font = GREEN; t[f'G{r}'].number_format = RUB; t[f'G{r}'].border = BORDER
        t[f'I{r}'] = f"=F{r}+G{r}*30"; t[f'I{r}'].font = BLACK; t[f'I{r}'].number_format = RUB; t[f'I{r}'].border = BORDER
    t[f'H{r}'] = "=Параметры!B5"; t[f'H{r}'].font = GREEN; t[f'H{r}'].number_format = RUB; t[f'H{r}'].border = BORDER
    r += 1
t['A4'].font = Font(name='Arial', italic=True)
# highlight ladder column D
for rr in range(4, 12):
    t[f'D{rr}'].fill = PatternFill('solid', start_color='FFF2CC')

# Fixed A->B mini-block
t['A14'] = 'Фиксированный тариф А→Б (на примере одной поездки)'; t['A14'].font = BOLD
def fb(r, label, formula, unit, fmt=RUB, font=BLACK, inp=False):
    t[f'A{r}'] = label; t[f'A{r}'].font = BLACK
    t[f'B{r}'] = formula; t[f'B{r}'].font = font; t[f'B{r}'].number_format = fmt; t[f'B{r}'].border = BORDER
    if inp:
        t[f'B{r}'].fill = YELLOW; t[f'B{r}'].font = BLUE
    t[f'C{r}'] = unit; t[f'C{r}'].font = Font(name='Arial', size=9, color='808080')
fb(15,'Прогноз пробега', 10, 'км', RUB, BLUE, True)
fb(16,'Прогноз длительности', 24, 'мин', RUB, BLUE, True)
fb(17,'Себестоимость поездки', "=(Параметры!B4/'Юнит-экономика'!B4)*B16+Параметры!B5*B15", 'руб', RUB, GREEN)
fb(18,'Цена по себестоимости с маржой', "=B17/(1-Параметры!B7)", 'руб', RUB)
fb(19,'Коммерческая скидка А→Б (стимул брать фикс)', 0.07, 'доля', PCT1, BLUE, True)
fb(20,'Итоговая цена А→Б', "=B18*(1-B19)", 'руб', RUB)
fb(21,'Та же поездка минутным тарифом', "=B16*'Юнит-экономика'!B16", 'руб', RUB, GREEN)
fb(22,'Выгода клиента к минутному', "=1-B20/B21", '%', PCT1)
t['B20'].fill = PatternFill('solid', start_color='FFF2CC')

t.column_dimensions['A'].width = 30
for col in 'BCDEFGHI':
    t.column_dimensions[col].width = 13

# ===================== СЕРВИСНЫЙ СБОР =====================
s = wb.create_sheet('Сервисный_сбор')
s['A1'] = 'Сервисный сбор: фикс. плата за поездку поверх тарифа'
s['A1'].font = TITLE
def srow(r, label, formula, fmt=RUB, font=BLACK, bold=False):
    s[f'A{r}'] = label; s[f'A{r}'].font = BOLD if bold else BLACK
    s[f'B{r}'] = formula; s[f'B{r}'].font = font; s[f'B{r}'].number_format = fmt; s[f'B{r}'].border = BORDER
srow(3,'Базовая выручка / сут (без сбора)', "='Юнит-экономика'!B11", RUB, GREEN)
srow(4,'Затраты / сут', "='Юнит-экономика'!B9", RUB, GREEN)
srow(5,'Поездок / сут', "=Параметры!B10", RUB, GREEN)
srow(6,'Средний чек поездки', "=B3/B5", RUB)

s['A8'] = 'Сценарии размера сбора'; s['A8'].font = BOLD
s['A9'] = 'Сбор за поездку, руб'; s['A9'].font = BOLDW; s['A9'].fill = HEADER; s['A9'].border = BORDER
for col, val in [('B',30),('C',50),('D',80)]:
    s[f'{col}9'] = val; s[f'{col}9'].font = BLUE; s[f'{col}9'].fill = YELLOW; s[f'{col}9'].number_format = RUB; s[f'{col}9'].border = BORDER

labels = [
 (10,'Доп. выручка / сут', lambda c: f"=$B$5*{c}9", RUB),
 (11,'Новая выручка / сут (ставки прежние)', lambda c: f"=$B$3+{c}10", RUB),
 (12,'Новая GP-маржа', lambda c: f"=($B$3+{c}10-$B$4)/($B$3+{c}10)", PCT1),
 (13,'Сбор как % среднего чека', lambda c: f"={c}9/$B$6", PCT1),
]
for r, lab, f, fmt in labels:
    s[f'A{r}'] = lab; s[f'A{r}'].font = BLACK
    for col in ('B','C','D'):
        s[f'{col}{r}'] = f(col); s[f'{col}{r}'].font = BLACK; s[f'{col}{r}'].number_format = fmt; s[f'{col}{r}'].border = BORDER

s['A15'] = 'Альтернатива: держим маржу 30%, снижаем минутную ставку'; s['A15'].font = BOLD
srow(16,'Целевая общая выручка (= затраты / (1-маржа))', "=$B$4/(1-Параметры!B7)", RUB, GREEN)
lab2 = [
 (17,'Нужная выручка с тарифа (минус сбор)', lambda c: f"=$B$16-{c}10"),
 (18,'Новая минутная ставка, руб/мин', lambda c: f"={c}17/'Юнит-экономика'!B4"),
]
for r, lab, f in lab2:
    s[f'A{r}'] = lab; s[f'A{r}'].font = BLACK
    for col in ('B','C','D'):
        s[f'{col}{r}'] = f(col); s[f'{col}{r}'].font = (GREEN if r==18 else BLACK); s[f'{col}{r}'].number_format = (RUB1 if r==18 else RUB); s[f'{col}{r}'].border = BORDER
srow(19,'Базовая минутная ставка (для сравнения)', "='Юнит-экономика'!B16", RUB1, GREEN)

s['A21'] = 'На что влияет сбор:'; s['A21'].font = BOLD
notes = [
 'сильнее всего бьёт по коротким поездкам (сбор = большая доля чека) → может срезать частые короткие сценарии',
 'двигает выручку из заметной минутной ставки в незаметный фикс → можно показывать ставку ниже конкурентов',
 'честно собирает фикс. косты на поездку (подача, уборка, поддержка, страховка), не зависящие от км/времени',
 'риски: падение конверсии в бронь, отток high-frequency юзеров; смягчается порогом/освобождением для подписки',
]
for i, n in enumerate(notes):
    s[f'A{22+i}'] = '•  ' + n; s[f'A{22+i}'].font = Font(name='Arial', size=10)

s.column_dimensions['A'].width = 52
for col in 'BCD':
    s.column_dimensions[col].width = 14

wb.save('carsharing_pricing.xlsx')
print('saved')
